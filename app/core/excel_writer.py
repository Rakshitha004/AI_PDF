import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def write_rows_to_excel(rows, output_path):
    df = pd.DataFrame(rows)
    df.index += 1
    df.index.name = "#"

    df.to_excel(output_path, engine="openpyxl")

    # Format after writing
    wb = load_workbook(output_path)
    ws = wb.active

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(output_path)



